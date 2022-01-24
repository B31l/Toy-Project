from openpyxl import Workbook, load_workbook
from tkinter import *
import tkinter.ttk
import os.path
PATH = "C:/WorkSpace/Toy-Project/PWstorage/"

root = Tk()
root.title("PWstorage")
root.geometry("800x400+200+100")
root.configure(bg="white")
root.resizable(False, False)

if not os.path.isfile(PATH + "data.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "DB"
    ws.append(["Name", "ID", "PW", "Description"])
    wb.save(PATH+"data.xlsx")
    wb.close() # wb 닫기
wb = load_workbook(PATH+"data.xlsx")
ws = wb.active
maxRow = ws.max_row # 데이터 수 + 1
NEWCOUNT = 0
DELCOUNT = 0

class File:
    def load():
        global databox
        databox = [[0,0,0,0] for i in range(maxRow-1+NEWCOUNT-DELCOUNT)]
        for x in range(maxRow-1+NEWCOUNT-DELCOUNT):
            for y in range(4):
                databox[x][y] = ws.cell(x+2, y+1).value
                if databox[x][y] == None : databox[x][y] = ""

    def save():
        wb.save(PATH+"data.xlsx")
    
class Data:
    def fill():
        File.load()
        treeview.delete(*treeview.get_children())
        Table.fill(databox)
        entry_Name.delete(0, END)
        entry_ID.delete(0, END)
        entry_PW.delete(0, END)
        entry_Description.delete(0, END)

    def new():
        global NEWCOUNT
        NEWCOUNT += 1
        ws.append([f"new cell {maxRow-1+NEWCOUNT-DELCOUNT}", "", "", ""])
        File.save()
        Data.fill()

    def delete():
        global DELCOUNT
        DELCOUNT += 1
        x = int(treeview.selection()[0])
        ws.delete_rows(x+2)
        File.save()
        Data.fill()

class DoubleClick:
    def edit(x):
        entry_Name.delete(0, END)
        entry_Name.insert(0, databox[x][0])
        entry_ID.delete(0, END)
        entry_ID.insert(0, databox[x][1])
        entry_PW.delete(0, END)
        entry_PW.insert(0, databox[x][2])
        entry_Description.delete(0, END)
        entry_Description.insert(0, databox[x][3])

    def edit_treeview(self):
        x= int(treeview.selection()[0])
        DoubleClick.edit(x)

    def edit_search(self):
        pass

class Table:
    def fill(box):
        for i in range(len(box)):
            treeview.insert("", "end", text=i+1, values=box[i], iid=str(i))

    def sort_Name():
        print("Name 정렬")

    def sort_ID():
        print("ID 정렬")

    def sort_PW():
        print("PW 정렬")

    def sort_Description():
        print("Description 정렬")

class Save:
    def click():
        x= int(treeview.selection()[0])
        ws.cell(row=x+2, column=1, value=entry_Name.get())
        ws.cell(row=x+2, column=2, value=entry_ID.get())    
        ws.cell(row=x+2, column=3, value=entry_PW.get())
        ws.cell(row=x+2, column=4, value=entry_Description.get())
        File.save()
        Data.fill()

    def enter(self):
        Save.click()

class Search:
    def click():
        listbox_Search.delete(0, END) # 검색창 초기화
        searchbox=[]
        resultbox=[]
        for i in range(len(databox)):
            searchbox.append(databox[i][0])
        result = entry_Search.get()

        for i in range(len(searchbox)): # 데이터 수만큼 반복
            if result == "":
                break
            count = len(searchbox[i])-len(result)+1
            for j in range(count): # 검색 결과와 데이터를 비교할 만큼 반복
                if searchbox[i][j:j + len(result)] == result: # 단어 길이만큼 일일이 비교
                    listbox_Search.insert(END, searchbox[i]) # 결과 내에 있으면 검색 결과에 출력
                    resultbox.append(databox[i])
                    continue
        treeview.delete(*treeview.get_children())
        Table.fill(resultbox)

    def enter(self):
        Search.click()


File.load()


frame_1 = Frame(root, bg="white", relief="flat", bd=1)
frame_1.grid(row=0, column=0, sticky=N+E+W+S)
btn_New = Button(frame_1, text="New", relief="solid", bg="white", padx=3, pady=3, width=8, height=1, command=Data.new)
btn_New.pack(side="left", fill="x", padx=5, pady=5)
btn_Delete = Button(frame_1, text="Delete", relief="solid", bg="white", padx=3, pady=3, width=8, command=Data.delete)
btn_Delete.pack(side="left", fill="x", padx=5, pady=5)
btn_Search = Button(frame_1, text="검색", relief="solid", bg="black", fg="white", width=8, command=Search.click)
btn_Search.pack(side="right", ipady=5, padx=5, pady=5)
entry_Search = Entry(frame_1, relief="solid", bd=2, width=35)
entry_Search.pack(side="right", ipady=5, pady=5)
entry_Search.bind("<Return>", Search.enter)


frame_2 = Frame(root, relief="solid", bd=1)
frame_2.grid(row=1, column=0, rowspan=2, sticky=N+E+W+S, padx=5, pady=5)
scrollbar = Scrollbar(frame_2)
scrollbar.pack(side="right", fill="y")
treeview = tkinter.ttk.Treeview\
    (frame_2, columns=["Name", "ID", "PW", "Description"],
    displaycolumns=["Name", "ID", "PW", "Description"],
    yscrollcommand=scrollbar.set,
    height=15)
treeview.pack(side="left")
treeview.column("#0", width=50)
treeview.heading("#0", command=Data.fill)
treeview.column("#1", width=100, anchor="w")
treeview.heading("Name", text="Name", command=Table.sort_Name)
treeview.column("#2", width=150, anchor="w")
treeview.heading("ID", text="ID", command=Table.sort_ID)
treeview.column("#3", width=30, anchor="w")
treeview.heading("PW", text="PW", command=Table.sort_PW)
treeview.column("#4", width=200, anchor="w")
treeview.heading("Description", text="Description", command=Table.sort_Description)
Table.fill(databox)
scrollbar.config(command=treeview.yview) # 매핑
treeview.bind("<Double-Button-1>", DoubleClick.edit_treeview)


frame_3 = Frame(root, bg="#33CC33", relief="solid", bd=1)
frame_3.grid(row=1, column=1, sticky=N+E+W+S, padx=5, pady=5)
Label(frame_3, text="이름", bg="#33CC33").grid(row=0, column=0, sticky=N+E+W+S, padx=3, pady=3)
Label(frame_3, text="ID", bg="#33CC33").grid(row=1, column=0, sticky=N+E+W+S, padx=3, pady=3)
Label(frame_3, text="PW", bg="#33CC33").grid(row=2, column=0, sticky=N+E+W+S, padx=3, pady=3)
Label(frame_3, text="설명", bg="#33CC33").grid(row=3, column=0, sticky=N+E+W+S, padx=3, pady=3)
entry_Name = Entry(frame_3, width=20)
entry_Name.grid(row=0, column=1, sticky=N+E+W+S, padx=3, pady=3)
entry_ID = Entry(frame_3, width=20)
entry_ID.grid(row=1, column=1, sticky=N+E+W+S, padx=3, pady=3)
entry_PW = Entry(frame_3, width=20)
entry_PW.grid(row=2, column=1, sticky=N+E+W+S, padx=3, pady=3)
entry_Description = Entry(frame_3, width=20)
entry_Description.grid(row=3, column=1, sticky=N+E+W+S, padx=3, pady=3)
btn_Save = Button\
    (frame_3, text="Save", relief="solid", bg="black", fg="white", padx=3, pady=3, width=3, command=Save.click)
btn_Save.grid(row=0, column=2, rowspan=4, sticky=N+E+W+S, padx=3, pady=3)
entry_Name.bind("<Return>", Save.enter)
entry_ID.bind("<Return>", Save.enter)
entry_PW.bind("<Return>", Save.enter)
entry_Description.bind("<Return>", Save.enter)


frame_4 = Frame(root, bg="white")
frame_4.grid(row=2, column=1, sticky=N+E+W+S, padx=5, pady=5)
listbox_Search = Listbox(frame_4, relief="solid", width=30, height=12)
listbox_Search.grid(row=0, column=0, sticky=N+E+W+S, padx=3, pady=3)
listbox_Search.bind("<Double-Button-1>", DoubleClick.edit_search)


root.mainloop()